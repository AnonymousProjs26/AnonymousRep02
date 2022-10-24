# -*- coding: UTF-8 -*-
import time
import os
import sys
import datetime
import numpy as np
import tensorflow.keras as keras
from tensorflow.keras.models import Model
import random
from tensorflow.keras.datasets import mnist, cifar10, cifar100
from numpy import arange
import openpyxl
import argparse
from tensorflow.keras.applications import vgg19, resnet50
from sklearn.cluster import KMeans
import math
import deepspeech_utils as ds_utils
from deepspeech_utils import DSDataUtils
from deepspeech_utils import dataGen_mfcc_ctc
import tensorflow.keras.backend as K
from deepspeech_text import Alphabet

def load_imagenet_result(model_name):
    filename1 = f"./imagenet_tools/dataset/{model_name}_train_50000.npz"
    train_result = np.load(filename1, allow_pickle=True)
    train_preds, train_labels = train_result["preds"], train_result["labels"]

    filename2 = f"./imagenet_tools/dataset/{model_name}_val_5000.npz"
    test_result = np.load(filename2, allow_pickle=True)
    test_preds, test_labels = test_result["preds"], test_result["labels"]
    return train_preds, train_labels, test_preds, test_labels

def get_model(exp_id):
    if exp_id == 'vgg19' or exp_id == 'vgg19_5000':
        my_model = vgg19.VGG19(weights='imagenet')
        adam = keras.optimizers.Adagrad(lr=0.01, epsilon=None, decay=0.0)
        my_model.compile(loss='categorical_crossentropy', optimizer=adam, metrics=['accuracy'])
    elif exp_id == 'resnet50' or exp_id == 'resnet50_5000':
        my_model = resnet50.ResNet50(weights='imagenet')
        adam = keras.optimizers.Adagrad(lr=0.01, epsilon=None, decay=0.0)
        my_model.compile(loss='categorical_crossentropy', optimizer=adam, metrics=['accuracy'])
    else:
        raise Exception("no such dataset found: {}".format(exp_id))

    return my_model

def get_acc(exp_id):
    acc_dict = {'vgg19': 0.7092,
                'resnet50': 0.7496,
                'vgg19_5000': 0.7146000266075134,
                'resnet50_5000': 0.7513999938964844}
    return acc_dict[exp_id]


if __name__ == "__main__":

    basedir = os.path.abspath(os.path.dirname(__file__))

    """Parser of command args"""
    parse = argparse.ArgumentParser()
    parse.add_argument("--exp_id", type=str, help="exp_identifiers")
    parse.add_argument("--cluster_alg", type=str, help="Clustering Algorithm")
    parse.add_argument("--cluster_num", type=int, help='Clustering parameter')
    # parse.add_argument("--pre_sample_num", type=int, help='Pre_sampling number')
    parse.add_argument("--random_seed", type=int, help='Random seed')

    console_flags, unparsed = parse.parse_known_args(sys.argv[1:])

    exp_id = console_flags.exp_id
    cluster_alg = console_flags.cluster_alg
    cluster_num = console_flags.cluster_num
    # pre_sample_num = console_flags.pre_sample_num
    random_seed = console_flags.random_seed

    acc = get_acc(exp_id)

    my_model = get_model(exp_id=exp_id)
    X_train_preds, Y_train, X_test_preds, Y_test = load_imagenet_result(model_name=exp_id)
    train_pred_labels = np.argmax(X_train_preds, axis=1)
    test_pred_labels = np.argmax(X_test_preds, axis=1)
    print(my_model.summary())

    start = datetime.datetime.now()

	dense_output = X_test_preds
	dense_output = np.max(dense_output, axis=1)
	dense_output = dense_output.reshape(-1, 1)
	print(dense_output)

    k_means = KMeans(n_clusters=cluster_num, random_state=10)

    k_means.fit(dense_output)

    print(k_means.cluster_centers_)
    print(k_means.inertia_)

    labels = k_means.labels_

    print(labels)
    print(np.max(labels))
    print(np.min(labels))

    y_pred_list = labels.tolist()
    countlist = []

    for i in range(np.min(labels), np.max(labels) + 1):
        countlist.append(y_pred_list.count(i))

    print(countlist)

    print(np.sort(countlist))
    print(np.argsort(countlist))

    res = {}
    for i, l in enumerate(labels):
        if l not in res:
            res[l] = []
        res[l].append(i)

    s_adaptive = {}
    acc_adaptive = {}
    total_ws = 0
    adaptive_index = {}

    total_adaptive_sampling = 0

    for key in res.keys():
        # index = res[key]
        # label = Y_test[index]
        # orig_sample = X_test[index]
        # pred = np.argmax(my_model.predict(orig_sample), axis=1)
        # label = np.argmax(label, axis=1)
        # s_stratum = np.std(pred == label)
        # print(key, "test_strata_Sh", s_stratum)
        # acc_stratum = np.sum(pred == label) / orig_sample.shape[0]
        # print(key, "test_strata_acc", acc_stratum)


        np.random.seed(random_seed)
        arr = np.random.permutation(res[key])
        random_index = arr[0:10]
        random_index = random_index.astype('int')

        total_adaptive_sampling = total_adaptive_sampling + len(random_index)

        adaptive_index[key] = random_index
        label = Y_test[random_index]
        orig_sample = X_test_preds[random_index]
        pred = np.argmax(orig_sample, axis=1)
        s_stratum = np.std(pred == label)
        print(key, "strata_Sh", s_stratum)
        print(pred)
        print(label)
        acc_stratum = np.sum(pred == label) / orig_sample.shape[0]
        print(key, "strata_acc", acc_stratum)
        s_adaptive[key] = s_stratum
        acc_adaptive[key] = acc_stratum

        total_ws = total_ws + len(res[key]) * s_stratum

    w = {}

    for key in res.keys():
        w[key] = len(res[key]) * s_adaptive[key] / total_ws

    print(w)
    # print(sum(w))

    delta = 10
    iterate = 98

    acc_select = []
    acc_random = []

    for k in range(50):
        print("the {} exp".format(k))

        acc_list1 = []
        acc_list2 = []
        statra_index1 = []

        acc_list3 = []

        random_index = []
        select_num = 50 - total_adaptive_sampling
        for i in range(iterate):
            test_selected_index = []
            acc_test = {}
            for key in res.keys():
                print(key, math.ceil(select_num * w[key]))
                if w[key] == 0 or math.ceil(select_num * w[key]) <= 0:
                    acc_1 = 1
                    acc_test[key] = acc_1
                else:
                    index = res[key]
                    index = np.array(index)
                    index = np.setdiff1d(index, adaptive_index[key], True)
                    arr = np.random.permutation(index.shape[0])
                    temp_index = arr[0:math.ceil(select_num * w[key])]
                    statra_index1 = index[temp_index]
                    statra_index1 = statra_index1.astype('int')

                    statra_index1 = np.union1d(statra_index1, adaptive_index[key])
                    print(key, len(statra_index1))
                    label = Y_test[statra_index1]
                    orig_sample = X_test_preds[statra_index1]
                    pred = np.argmax(orig_sample, axis=1)
                    acc_1 = np.sum(pred == label) / orig_sample.shape[0]
                    acc_test[key] = acc_1

                    test_selected_index.extend(statra_index1)

            arr = np.random.permutation(X_test_preds.shape[0])
            random_index = arr[0:select_num]
            random_index = random_index.astype('int')

            acc1 = 0
            print(acc_test)
            for key in res.keys():
                acc1 = acc1 + acc_test[key] * (len(res[key]) / len(X_test_preds))

            acc_list1.append(acc1)
            label = Y_test[random_index]
            orig_sample = X_test_preds[random_index]
            pred = np.argmax(orig_sample, axis=1)
            acc2 = np.sum(pred == label) / orig_sample.shape[0]
            acc_list2.append(acc2)

            print("numuber of samples is {!s}, select acc is {!s}, random acc is {!s}".format(
                select_num+total_adaptive_sampling, acc1, acc2))
            print(np.abs(acc1 - acc), np.abs(acc2 - acc))

            select_num = select_num + delta

        acc_list1 = np.square(np.array(acc_list1) - acc)
        acc_list2 = np.square(np.array(acc_list2) - acc)

        acc_select.append(acc_list1)
        acc_random.append(acc_list2)

        elapsed = (datetime.datetime.now() - start)
        print("Time used: ", elapsed)

    acc_select = np.array(acc_select)
    acc_random = np.array(acc_random)

    mse_select = np.sqrt(acc_select.mean(axis=0))
    mse_random = np.sqrt(acc_random.mean(axis=0))

    # save the result
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    row_num = 3
    sheet.cell(row=row_num, column=1).value = "SSOA-C-P"
    for i in range(len(mse_select)):
        sheet.cell(row=row_num, column=i + 2).value = mse_select[i]
    row_num += 1
    for i in range(len(mse_random)):
        sheet.cell(row=row_num, column=i + 2).value = mse_random[i]

    workbook.save(os.path.join(basedir, "results", "{}-ssoacp-{}-{}-seed{}.xlsx".format(exp_id, cluster_alg, cluster_num, random_seed)))

    elapsed = (datetime.datetime.now() - start)
    print("Time used: ", elapsed)
