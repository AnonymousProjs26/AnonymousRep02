# -*- coding: UTF-8 -*-
import time
import os
import sys
import datetime
import numpy as np
import tensorflow.keras as keras
from tensorflow.keras.models import Model
import random
from tensorflow.keras.datasets import mnist, cifar10, cifar100
from numpy import arange
import hdbscan
import openpyxl
import argparse
from tensorflow.keras.applications import vgg19, resnet50
from sklearn.cluster import KMeans
import math

def get_score(x_test, y_test, model):
    score = model.evaluate(x_test, y_test, verbose=0)
    print('Test loss:', score[0])
    print('Test accuracy:', score[1])
    return score

def get_cifar10_cn12(**kwargs):
    (X_train, Y_train), (X_test, Y_test) = cifar10.load_data()
    X_test = X_test.astype('float32')
    X_test = (X_test / 255.0) - (1.0 - 0.5)
    Y_test = keras.utils.to_categorical(Y_test, 10)

    X_train = X_train.astype('float32')
    X_train = (X_train / 255.0) - (1.0 - 0.5)
    Y_train = keras.utils.to_categorical(Y_train, 10)
    return X_test, Y_test, X_train, Y_train


def get_cifar10_vgg16(**kwargs):
    (X_train, Y_train), (X_test, Y_test) = cifar10.load_data()

    X_train = X_train.astype('float32')
    X_test = X_test.astype('float32')
    mean = np.mean(X_train, axis=(0, 1, 2, 3))
    std = np.std(X_train, axis=(0, 1, 2, 3))
    
    X_train = (X_train - mean) / (std + 1e-7)
    X_test = (X_test - mean) / (std + 1e-7)

    Y_test = keras.utils.to_categorical(Y_test, 10)
    Y_train = keras.utils.to_categorical(Y_train, 10)
    return X_test, Y_test, X_train, Y_train


def get_cifar100_vgg16(**kwargs):
    (X_train, Y_train), (X_test, Y_test) = cifar100.load_data()

    X_train = X_train.astype('float32')
    X_test = X_test.astype('float32')
    mean = np.mean(X_train, axis=(0, 1, 2, 3))
    std = np.std(X_train, axis=(0, 1, 2, 3))
    
    X_train = (X_train - mean) / (std + 1e-7)
    X_test = (X_test - mean) / (std + 1e-7)

    Y_test = keras.utils.to_categorical(Y_test, 100)
    Y_train = keras.utils.to_categorical(Y_train, 100)
    return X_test, Y_test, X_train, Y_train

def get_imagenet(**kwargs):
    data_path = os.path.join(basedir, 'data', "imagenet.npz")
    data = np.load(data_path)
    X_test, Y_test = data['x_test'], data['y_test']
    exp_id = kwargs['exp_id']
    if exp_id == 'vgg19':
        X_test = vgg19.preprocess_input(X_test)
        Y_test = keras.utils.to_categorical(Y_test, num_classes=1000)
    if exp_id == 'resnet50':
        X_test = resnet50.preprocess_input(X_test)
        Y_test = keras.utils.to_categorical(Y_test, num_classes=1000)
    return X_test, Y_test

def get_data(exp_id):
    exp_model_dict = {'cn12': get_cifar10_cn12,
                      'cifar10_vgg16': get_cifar10_vgg16,
                      'cifar100_vgg16': get_cifar100_vgg16}
    return exp_model_dict[exp_id](exp_id=exp_id)


def get_model(exp_id):
    basedir = os.path.abspath(os.path.dirname(__file__))

    exp_model_dict = {'cn12': 'model/cifar10_cn12.h5',
                      'cifar10_vgg16': 'model/cifar10vgg.h5',
                      'cifar100_vgg16': 'model/cifar100vgg.h5'}

    if exp_id == 'vgg19':
        my_model = vgg19.VGG19(weights='imagenet')
        adam = keras.optimizers.Adagrad(lr=0.01, epsilon=None, decay=0.0)
        my_model.compile(loss='categorical_crossentropy', optimizer=adam, metrics=['accuracy'])
    elif exp_id == 'resnet50':
        my_model = resnet50.ResNet50(weights='imagenet')
        adam = keras.optimizers.Adagrad(lr=0.01, epsilon=None, decay=0.0)
        my_model.compile(loss='categorical_crossentropy', optimizer=adam, metrics=['accuracy'])
    elif exp_id in exp_model_dict.keys():
        my_model = keras.models.load_model(os.path.join(basedir, exp_model_dict[exp_id]))
    else:
        raise Exception("no such dataset found: {}".format(exp_id))

    return my_model


def get_acc(exp_id):
    acc_dict = {'cn12': 0.8064,
                'cifar10_vgg16': 0.9375,
                'cifar100_vgg16': 0.6944}
    return acc_dict[exp_id]


def get_cluster_vec(model, X):
    dense_output = np.max(model.predict(X), axis=1)
    dense_output = dense_output.reshape(-1, 1)
    print(dense_output)
    return dense_output


def cluster_by_kmeans(cluster_num, test_vec, train_vec):
    k_means = KMeans(n_clusters=cluster_num, random_state=10)

    k_means.fit(test_vec)

    print(k_means.cluster_centers_)
    print(k_means.inertia_)

    labels = k_means.labels_

    print(labels)
    print(np.max(labels))
    print(np.min(labels))

    y_pred_list = labels.tolist()
    countlist = []

    for i in range(np.min(labels), np.max(labels) + 1):
        countlist.append(y_pred_list.count(i))

    print(countlist)

    print(np.sort(countlist))
    print(np.argsort(countlist))

    res = {}
    for i, l in enumerate(labels):
        if l not in res:
            res[l] = []
        res[l].append(i)

    for key in res:
        X_test3 = []
        Y_test3 = []
        print(key, len(res[key]))
        for i in range(len(res[key])):
            X_test3.append(X_test[res[key][i]])
            Y_test3.append(Y_test[res[key][i]])
        X_test3 = np.array(X_test3)
        Y_test3 = np.array(Y_test3)
        score = get_score(X_test3, Y_test3, my_model)

    res_train = {}

    for i in range(len(train_vec)):
        temp = []
        for center in k_means.cluster_centers_:
            temp.append(math.sqrt(np.power(center - dense_output_train[i], 2).sum()))
        temp = np.array(temp)
        # print(np.argsort(temp))
        k_train = np.argsort(temp)[0]
        if k_train not in res_train:
            res_train[k_train] = []
        res_train[k_train].append(i)
        
        if i % 1000 == 0:
            print(i, res_train.keys())

    for key in res_train:
        X_test3 = []
        Y_test3 = []
        print(key, len(res_train[key]))
        for i in range(len(res_train[key])):
            X_test3.append(X_train[res_train[key][i]])
            Y_test3.append(Y_train[res_train[key][i]])
        X_test3 = np.array(X_test3)
        Y_test3 = np.array(Y_test3)
        score = get_score(X_test3, Y_test3, my_model)
    return res, res_train

if __name__ == "__main__":

    basedir = os.path.abspath(os.path.dirname(__file__))

    """Parser of command args"""
    parse = argparse.ArgumentParser()
    parse.add_argument("--exp_id", type=str, help="exp_identifiers")
    parse.add_argument("--cluster_alg", type=str, help="Clustering Algorithm")
    parse.add_argument("--cluster_num", type=int, help='Clustering parameter')

    console_flags, unparsed = parse.parse_known_args(sys.argv[1:])

    exp_id = console_flags.exp_id
    cluster_alg = console_flags.cluster_alg
    cluster_num = console_flags.cluster_num

    acc = get_acc(exp_id)

    my_model = get_model(exp_id=exp_id)
    X_test, Y_test, X_train, Y_train = get_data(exp_id=exp_id)
    print(my_model.summary())
    print(X_test.shape)
    print(X_train.shape)

    # get_score(X_test,Y_test,my_model)
    # get_score(X_train,Y_train,my_model)

    start = datetime.datetime.now()

    dense_output_test = get_cluster_vec(my_model, X_test)
    dense_output_train = get_cluster_vec(my_model, X_train)

    if cluster_alg == 'kmeans':
        res, res_train = cluster_by_kmeans(cluster_num, dense_output_test, dense_output_train)

    s_train = {}
    acc_train = {}
    len_index_train = {}
    total_ws = 0

    for key in res.keys():
        if key in res_train.keys() and len(res_train[key]) >= 10:
            index = res_train[key]
            print(len(index))
            len_index_train[key] = len(res[key])
            label = Y_train[index]
            orig_sample = X_train[index]
            pred = np.argmax(my_model.predict(orig_sample), axis=1)
            label = np.argmax(label, axis=1)
            s_stratum = np.std(pred == label)
            print(key, "strata_Sh", s_stratum)
            acc_stratum = np.sum(pred == label) / orig_sample.shape[0]
            print(key, "strata_acc", acc_stratum)
            s_train[key] = s_stratum
            acc_train[key] = acc_stratum

            total_ws = total_ws + len_index_train[key] * s_stratum
        else:
            arr = np.random.permutation(res[key])
            random_index = arr[0:10]
            random_index = random_index.astype('int')

            len_index_train[key] = len(res[key])

            label = Y_test[random_index]
            orig_sample = X_test[random_index]
            pred = np.argmax(my_model.predict(orig_sample), axis=1)
            label = np.argmax(label, axis=1)
            s_stratum = np.std(pred == label)
            print(key, "strata_Sh", s_stratum)
            acc_stratum = np.sum(pred == label) / orig_sample.shape[0]
            print(key, "strata_acc", acc_stratum)
            s_train[key] = s_stratum
            acc_train[key] = acc_stratum

            total_ws = total_ws + len(res[key]) * s_stratum

    w = {}

    for key in res.keys():
        w[key] = len_index_train[key] * s_train[key] / total_ws

    print(w)
    # print(sum(w))

    delta = 10
    iterate = 17

    acc_select = []
    acc_random = []

    for k in range(50):
        print("the {} exp".format(k))

        acc_list1 = []
        acc_list2 = []
        statra_index1 = []

        random_index = []
        select_num = 50
        for i in range(iterate):
            test_selected_index = []
            acc_test = {}
            for key in res.keys():
                print(key, math.ceil(select_num * w[key]))
                if w[key] == 0 or math.ceil(select_num * w[key]) <= 0:
                    acc_1 = 1
                    acc_test[key] = acc_1
                else:
                    index = res[key]
                    index = np.array(index)
                    arr = np.random.permutation(index.shape[0])
                    temp_index = arr[0:math.ceil(select_num * w[key])]
                    statra_index1 = index[temp_index]
                    statra_index1 = statra_index1.astype('int')
                    label = Y_test[statra_index1]
                    orig_sample = X_test[statra_index1]
                    pred = np.argmax(my_model.predict(orig_sample), axis=1)
                    label = np.argmax(label, axis=1)
                    acc_1 = np.sum(pred == label) / orig_sample.shape[0]
                    acc_test[key] = acc_1

                    test_selected_index.extend(statra_index1)

            arr = np.random.permutation(X_test.shape[0])
            random_index = arr[0:select_num]
            random_index = random_index.astype('int')

            acc1 = 0
            print(acc_test)
            for key in res.keys():
                acc1 = acc1 + acc_test[key] * (len(res[key]) / len(X_test))

            acc_list1.append(acc1)
            label = Y_test[random_index]
            orig_sample = X_test[random_index]
            pred = np.argmax(my_model.predict(orig_sample), axis=1)
            label = np.argmax(label, axis=1)
            acc2 = np.sum(pred == label) / orig_sample.shape[0]
            acc_list2.append(acc2)

            print("numuber of samples is {!s}, select acc is {!s}, random acc is {!s}".format(
                select_num, acc1, acc2))
            print(np.abs(acc1 - acc), np.abs(acc2 - acc))

            select_num = select_num + delta

        acc_list1 = np.square(np.array(acc_list1) - acc)
        acc_list2 = np.square(np.array(acc_list2) - acc)

        acc_select.append(acc_list1)
        acc_random.append(acc_list2)

        elapsed = (datetime.datetime.now() - start)
        print("Time used: ", elapsed)

    acc_select = np.array(acc_select)
    acc_random = np.array(acc_random)

    mse_select = np.sqrt(acc_select.mean(axis=0))
    mse_random = np.sqrt(acc_random.mean(axis=0))

    # save the result
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    row_num = 5

    sheet.cell(row=row_num, column=1).value = "SSOA-C-T"
    for i in range(len(mse_select)):
        sheet.cell(row=row_num, column=i + 2).value = mse_select[i]
    row_num += 1
    for i in range(len(mse_random)):
        sheet.cell(row=row_num, column=i + 2).value = mse_random[i]

    workbook.save(os.path.join(basedir, "results", "{}-ssoact-{}-{}.xlsx".format(exp_id, cluster_alg, cluster_num)))

    elapsed = (datetime.datetime.now() - start)
    print("Time used: ", elapsed)
