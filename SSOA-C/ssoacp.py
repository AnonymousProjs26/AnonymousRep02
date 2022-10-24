# -*- coding: UTF-8 -*-
import time
import os
import sys
import datetime
import numpy as np
import tensorflow.keras as keras
from tensorflow.keras.models import Model
import random
from tensorflow.keras.datasets import mnist, cifar10, cifar100
from numpy import arange
import openpyxl
import argparse
from tensorflow.keras.applications import vgg19, resnet50
from sklearn.cluster import KMeans
import math
import deepspeech_utils as ds_utils
from deepspeech_utils import DSDataUtils
from deepspeech_utils import dataGen_mfcc_ctc
import tensorflow.keras.backend as K
from deepspeech_text import Alphabet

def get_score(x_test, y_test, model):
    score = model.evaluate(x_test, y_test, verbose=0)
    print('Test loss:', score[0])
    print('Test accuracy:', score[1])
    return score

def get_cifar10_cn12(**kwargs):
    (X_train, Y_train), (X_test, Y_test) = cifar10.load_data()
    X_test = X_test.astype('float32')
    X_test = (X_test / 255.0) - (1.0 - 0.5)
    Y_test = keras.utils.to_categorical(Y_test, 10)

    X_train = X_train.astype('float32')
    X_train = (X_train / 255.0) - (1.0 - 0.5)
    Y_train = keras.utils.to_categorical(Y_train, 10)
    return X_test, Y_test, X_train, Y_train

def get_cifar10_vgg16(**kwargs):
    (X_train, Y_train), (X_test, Y_test) = cifar10.load_data()

    X_train = X_train.astype('float32')
    X_test = X_test.astype('float32')
    mean = np.mean(X_train, axis=(0, 1, 2, 3))
    std = np.std(X_train, axis=(0, 1, 2, 3))
    
    X_train = (X_train - mean) / (std + 1e-7)
    X_test = (X_test - mean) / (std + 1e-7)

    Y_test = keras.utils.to_categorical(Y_test, 10)
    Y_train = keras.utils.to_categorical(Y_train, 10)
    return X_test, Y_test, X_train, Y_train

def get_cifar100_vgg16(**kwargs):
    (X_train, Y_train), (X_test, Y_test) = cifar100.load_data()

    X_train = X_train.astype('float32')
    X_test = X_test.astype('float32')
    mean = np.mean(X_train, axis=(0, 1, 2, 3))
    std = np.std(X_train, axis=(0, 1, 2, 3))
    
    X_train = (X_train - mean) / (std + 1e-7)
    X_test = (X_test - mean) / (std + 1e-7)

    Y_test = keras.utils.to_categorical(Y_test, 100)
    Y_train = keras.utils.to_categorical(Y_train, 100)
    return X_test, Y_test, X_train, Y_train

def get_combined_cifar10(**kwargs):
    # Load the CIFAR10 data.
    (x_train, y_train), (x_test, y_test) = cifar10.load_data()

    x_train = x_train.astype('float32')
    x_test = x_test.astype('float32')
    x_train /= 255
    x_test /= 255

    basedir = os.path.abspath(os.path.dirname(__file__))

    adv_data_path = os.path.join(basedir, 'data/cifar10_vgg16_jsma_full.npz')
    # label_path = os.path.join(basedir,'data','adv_image/bim_mnist_label.npy')
    adv_data = np.load(adv_data_path)

    x_adv = adv_data["inputs"]
    y_adv = adv_data["labels"]

    y_adv = y_adv.squeeze()

    x_test = x_test[:5000]
    y_test = y_test[:5000]

    y_test = y_test.reshape(5000)

    x_test = np.concatenate((x_test, x_adv[:5000]), axis=0)
    y_test = np.concatenate((y_test, y_adv[:5000]), axis=0)

    # Convert class vectors to binary class matrices.
    y_train = keras.utils.to_categorical(y_train, 10)
    # y_train = np.argmax(y_train, axis=1)
    y_test = keras.utils.to_categorical(y_test, 10)
    # y_test = np.argmax(y_test, axis=1)

    print(y_train)
    print(y_test)

    print(x_train.shape)
    print(y_train.shape)

    print(x_test.shape)
    print(y_test.shape)

    return x_test, y_test, x_train, y_train


def subtract_mean(x, means):
    x[..., 0] -= means[0]
    x[..., 1] -= means[1]
    x[..., 2] -= means[2]
    return x


def get_combined_cifar100(**kwargs):
    (x_train, y_train), (x_test, y_test) = cifar100.load_data()

    x_train = x_train.astype('float32')
    x_test = x_test.astype('float32')
    x_train /= 255
    x_test /= 255

    means = [0.5070746, 0.48654896, 0.44091788]

    x_train = subtract_mean(x_train, means)
    x_test = subtract_mean(x_test, means)

    basedir = os.path.abspath(os.path.dirname(__file__))

    adv_data_path = os.path.join(basedir, 'data/cifar100_resnet32_fgsm_full.npz')
    # label_path = os.path.join(basedir,'data','adv_image/bim_mnist_label.npy')
    adv_data = np.load(adv_data_path)

    x_adv = adv_data["inputs"]
    y_adv = adv_data["labels"]

    y_adv = y_adv.squeeze()

    x_test = x_test[:5000]
    y_test = y_test[:5000]

    y_test = y_test.reshape(5000)

    x_test = np.concatenate((x_test, x_adv[:5000]), axis=0)
    y_test = np.concatenate((y_test, y_adv[:5000]), axis=0)

    # Convert class vectors to binary class matrices.
    y_train = keras.utils.to_categorical(y_train, 100)
    # y_train = np.argmax(y_train, axis=1)
    y_test = keras.utils.to_categorical(y_test, 100)
    # y_test = np.argmax(y_test, axis=1)

    print(y_train)
    print(y_test)

    print(x_train.shape)
    print(y_train.shape)

    print(x_test.shape)
    print(y_test.shape)

    return x_test, y_test, x_train, y_train


def get_data(exp_id):
    exp_model_dict = {'combined_cifar10': get_combined_cifar10,
                      'combined_cifar10': get_combined_cifar10,
                      'cn12': get_cifar10_cn12,
                      'cifar10_vgg16': get_cifar10_vgg16,
                      'cifar100_vgg16': get_cifar100_vgg16}
    return exp_model_dict[exp_id](exp_id=exp_id)


def get_model(exp_id):
    basedir = os.path.abspath(os.path.dirname(__file__))

    exp_model_dict = {'combined_cifar10': 'model/cifar10_vgg16.h5',
                      'combined_cifar100': 'model/cifar100_resnet32.h5',
                      'cn12': 'model/cifar10_cn12.h5',
                      'cifar10_vgg16': 'model/cifar10vgg.h5',
                      'cifar100_vgg16': 'model/cifar100vgg.h5'}

    if exp_id == 'vgg19' or exp_id == 'vgg19_5000':
        my_model = vgg19.VGG19(weights='imagenet')
        adam = keras.optimizers.Adagrad(lr=0.01, epsilon=None, decay=0.0)
        my_model.compile(loss='categorical_crossentropy', optimizer=adam, metrics=['accuracy'])
    elif exp_id == 'resnet50' or exp_id == 'resnet50_5000':
        my_model = resnet50.ResNet50(weights='imagenet')
        adam = keras.optimizers.Adagrad(lr=0.01, epsilon=None, decay=0.0)
        my_model.compile(loss='categorical_crossentropy', optimizer=adam, metrics=['accuracy'])
    elif exp_id in exp_model_dict.keys():
        my_model = keras.models.load_model(os.path.join(basedir, exp_model_dict[exp_id]))
    else:
        raise Exception("no such dataset found: {}".format(exp_id))

    return my_model


def get_acc(exp_id):
    acc_dict = {'combined_cifar10': 0.437,
                'combined_cifar100': 0.3504,
                'cn12': 0.8064,
                'cifar10_vgg16': 0.9375,
                'cifar100_vgg16': 0.6944}
    return acc_dict[exp_id]


if __name__ == "__main__":

    basedir = os.path.abspath(os.path.dirname(__file__))

    """Parser of command args"""
    parse = argparse.ArgumentParser()
    parse.add_argument("--exp_id", type=str, help="exp_identifiers")
    parse.add_argument("--cluster_alg", type=str, help="Clustering Algorithm")
    parse.add_argument("--cluster_num", type=int, help='Clustering parameter')
    parse.add_argument("--random_seed", type=int, help='Random seed')

    console_flags, unparsed = parse.parse_known_args(sys.argv[1:])

    exp_id = console_flags.exp_id
    cluster_alg = console_flags.cluster_alg
    cluster_num = console_flags.cluster_num
    random_seed = console_flags.random_seed

    acc = get_acc(exp_id)

    my_model = get_model(exp_id=exp_id)
    X_test,Y_test,X_train,Y_train = get_data(exp_id=exp_id)
    # X_test, Y_test = get_data(exp_id=exp_id)
    print(my_model.summary())

    # get_score(X_test, Y_test, my_model)
    # get_score(X_train,Y_train,my_model)

    start = datetime.datetime.now()

	dense_output = my_model.predict(X_test)
	dense_output = np.max(dense_output, axis=1)
	dense_output = dense_output.reshape(-1, 1)
	print(dense_output)

    k_means = KMeans(n_clusters=cluster_num, random_state=10)

    k_means.fit(dense_output)

    print(k_means.cluster_centers_)
    print(k_means.inertia_)

    labels = k_means.labels_

    print(labels)
    print(np.max(labels))
    print(np.min(labels))

    y_pred_list = labels.tolist()
    countlist = []

    for i in range(np.min(labels), np.max(labels) + 1):
        countlist.append(y_pred_list.count(i))

    print(countlist)

    print(np.sort(countlist))
    print(np.argsort(countlist))

    res = {}
    for i, l in enumerate(labels):
        if l not in res:
            res[l] = []
        res[l].append(i)

    for key in res:
        X_test3 = []
        Y_test3 = []
        print(key, len(res[key]))
        for i in range(len(res[key])):
            X_test3.append(X_test[res[key][i]])
            Y_test3.append(Y_test[res[key][i]])
        X_test3 = np.array(X_test3)
        Y_test3 = np.array(Y_test3)
        score = get_score(X_test3, Y_test3, my_model)

    s_adaptive = {}
    acc_adaptive = {}
    # len_index_adaptive = {}
    total_ws = 0
    adaptive_index = {}

    total_adaptive_sampling = 0

    for key in res.keys():
        # index = res[key]
        # label = Y_test[index]
        # orig_sample = X_test[index]
        # pred = np.argmax(my_model.predict(orig_sample), axis=1)
        # label = np.argmax(label, axis=1)
        # s_stratum = np.std(pred == label)
        # print(key, "test_strata_Sh", s_stratum)
        # acc_stratum = np.sum(pred == label) / orig_sample.shape[0]
        # print(key, "test_strata_acc", acc_stratum)

        np.random.seed(random_seed)
        arr = np.random.permutation(res[key])
        random_index = arr[0:10]
        random_index = random_index.astype('int')

        total_adaptive_sampling = total_adaptive_sampling + len(random_index)

        adaptive_index[key] = random_index
        label = Y_test[random_index]
        orig_sample = X_test[random_index]
        pred = np.argmax(my_model.predict(orig_sample), axis=1)
        label = np.argmax(label, axis=1)
        s_stratum = np.std(pred == label)
        print(key, "strata_Sh", s_stratum)
        acc_stratum = np.sum(pred == label) / orig_sample.shape[0]
        print(key, "strata_acc", acc_stratum)
        s_adaptive[key] = s_stratum
        acc_adaptive[key] = acc_stratum

        total_ws = total_ws + len(res[key]) * s_stratum

    w = {}

    for key in res.keys():
        w[key] = len(res[key]) * s_adaptive[key] / total_ws

    print(w)
    # print(sum(w))

    delta = 10
    iterate = 100

    acc_select = []
    acc_random = []

    for k in range(50):
        print("the {} exp".format(k))

        acc_list1 = []
        acc_list2 = []
        statra_index1 = []

        acc_list3 = []

        random_index = []
        select_num = 50 - total_adaptive_sampling
        for i in range(iterate):
            test_selected_index = []
            acc_test = {}
            for key in res.keys():
                print(key, math.ceil(select_num * w[key]))
                if w[key] == 0 or math.ceil(select_num * w[key]) <= 0:
                    acc_1 = 1
                    acc_test[key] = acc_1
                else:
                    index = res[key]
                    index = np.array(index)
                    index = np.setdiff1d(index, adaptive_index[key], True)
                    arr = np.random.permutation(index.shape[0])
                    temp_index = arr[0:math.ceil(select_num * w[key])]
                    statra_index1 = index[temp_index]
                    statra_index1 = statra_index1.astype('int')

                    statra_index1 = np.union1d(statra_index1, adaptive_index[key])
                    print(key, len(statra_index1))
                    label = Y_test[statra_index1]
                    orig_sample = X_test[statra_index1]
                    pred = np.argmax(my_model.predict(orig_sample), axis=1)
                    label = np.argmax(label, axis=1)
                    acc_1 = np.sum(pred == label) / orig_sample.shape[0]
                    acc_test[key] = acc_1

                    test_selected_index.extend(statra_index1)

            arr = np.random.permutation(X_test.shape[0])
            random_index = arr[0:select_num]
            random_index = random_index.astype('int')

            acc1 = 0
            print(acc_test)
            for key in res.keys():
                acc1 = acc1 + acc_test[key] * (len(res[key]) / len(X_test))

            acc_list1.append(acc1)
            label = Y_test[random_index]
            orig_sample = X_test[random_index]
            pred = np.argmax(my_model.predict(orig_sample), axis=1)
            label = np.argmax(label, axis=1)
            acc2 = np.sum(pred == label) / orig_sample.shape[0]
            acc_list2.append(acc2)

            print("numuber of samples is {!s}, select acc is {!s}, random acc is {!s}".format(
                select_num+total_adaptive_sampling, acc1, acc2))
            print(np.abs(acc1 - acc), np.abs(acc2 - acc))

            select_num = select_num + delta

        acc_list1 = np.square(np.array(acc_list1) - acc)
        acc_list2 = np.square(np.array(acc_list2) - acc)

        acc_select.append(acc_list1)
        acc_random.append(acc_list2)

        elapsed = (datetime.datetime.now() - start)
        print("Time used: ", elapsed)

    acc_select = np.array(acc_select)
    acc_random = np.array(acc_random)

    mse_select = np.sqrt(acc_select.mean(axis=0))
    mse_random = np.sqrt(acc_random.mean(axis=0))

    # save the result
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    row_num = 3
    sheet.cell(row=row_num, column=1).value = "SSOA-C-P"
    for i in range(len(mse_select)):
        sheet.cell(row=row_num, column=i + 2).value = mse_select[i]
    row_num += 1
    for i in range(len(mse_random)):
        sheet.cell(row=row_num, column=i + 2).value = mse_random[i]

    workbook.save(os.path.join(basedir, "results", "{}-ssoacp-{}-{}.xlsx".format(exp_id, cluster_alg, cluster_num)))

    elapsed = (datetime.datetime.now() - start)
    print("Time used: ", elapsed)
