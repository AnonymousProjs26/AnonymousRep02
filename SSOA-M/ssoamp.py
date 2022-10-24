# -*- coding: UTF-8 -*-
import sys
import datetime
import os
import numpy as np
import tensorflow.keras.backend as K
from collections import defaultdict
import matplotlib.pyplot as plt
import tensorflow.keras as keras
from sklearn.metrics import *
import math
from tensorflow.keras.datasets import cifar10, cifar100
from tensorflow.keras.models import load_model
import argparse
from tensorflow.keras.applications import vgg19, resnet50
import openpyxl

def get_score(x_test, y_test, model):
    score = model.evaluate(x_test, y_test, verbose=0)
    print('Test loss:', score[0])
    print('Test accuracy:', score[1])
    return score

def selectsample(model, x_test, y_test, delta, iterate, fig_size):
    global index3, index1, index2, index1_adaptive, index2_adaptive, index3_adaptive, total_adaptive_sampling, w
    acc_list1 = []
    acc_list2 = []

    select_num = 50 - total_adaptive_sampling
    for i in range(iterate):
        if w[0] == 0 or math.ceil(select_num * w[0]) < 1:
            acc_1 = 1
        else:
            index1 = np.setdiff1d(index1, index1_adaptive, True)
            arr = np.random.permutation(index1.shape[0])
            temp_index = arr[0:math.ceil(select_num * w[0])]
            statra_index1 = index1[temp_index]
            statra_index1 = statra_index1.astype('int')

            statra_index1 = np.union1d(statra_index1, index1_adaptive)
            print("0: ", len(statra_index1))
            label = y_test[statra_index1]
            orig_sample = x_test[statra_index1]
            orig_sample = orig_sample.reshape(-1, fig_size, fig_size, 3)
            pred = np.argmax(model.predict(orig_sample), axis=1)
            acc_1 = np.sum(pred == label) / orig_sample.shape[0]


        if w[1] == 0 or math.ceil(select_num * w[1]) < 1:
            acc_2 = 1
        else:
            index2 = np.setdiff1d(index2, index2_adaptive, True)
            arr = np.random.permutation(index2.shape[0])
            temp_index = arr[0:math.ceil(select_num * w[1])]
            statra_index2 = index2[temp_index]
            statra_index2 = statra_index2.astype('int')

            statra_index2 = np.union1d(statra_index2, index2_adaptive)
            print("1: ", len(statra_index2))
            label = y_test[statra_index2]
            orig_sample = x_test[statra_index2]
            orig_sample = orig_sample.reshape(-1, fig_size, fig_size, 3)
            pred = np.argmax(model.predict(orig_sample), axis=1)
            acc_2 = np.sum(pred == label) / orig_sample.shape[0]

        if w[2] == 0 or math.ceil(select_num * w[2]) < 1:
            acc_3 = 1
        else:
            index3 = np.setdiff1d(index3, index3_adaptive, True)
            arr = np.random.permutation(index3.shape[0])
            temp_index = arr[0:math.ceil(select_num * w[2])]
            statra_index3 = index3[temp_index]
            statra_index3 = statra_index3.astype('int')

            statra_index3 = np.union1d(statra_index3, index3_adaptive)
            print("2: ", len(statra_index3))
            label = y_test[statra_index3]
            orig_sample = x_test[statra_index3]
            orig_sample = orig_sample.reshape(-1, fig_size, fig_size, 3)
            pred = np.argmax(model.predict(orig_sample), axis=1)
            acc_3 = np.sum(pred == label) / orig_sample.shape[0]

        arr = np.random.permutation(x_test.shape[0])
        random_index = arr[0:select_num+total_adaptive_sampling]
        random_index = random_index.astype('int')

        acc1 = 0.1 * acc_1 + 0.1 * acc_2 + 0.8 * acc_3
        acc_list1.append(acc1)

        label = y_test[random_index]
        orig_sample = x_test[random_index]
        orig_sample = orig_sample.reshape(-1, fig_size, fig_size, 3)
        pred = np.argmax(model.predict(orig_sample), axis=1)
        acc2 = np.sum(pred == label) / orig_sample.shape[0]
        acc_list2.append(acc2)

        print("numuber of samples is {!s}, select acc is {!s}, random acc is {!s}".format(
            orig_sample.shape[0], acc1, acc2))

        select_num = select_num + delta

    return acc_list1, acc_list2


def experiments(delta, iterate, fig_size):
    pred = np.argmax(model.predict(x_test), axis=1)
    true_acc = np.sum(pred == y_test) / x_test.shape[0]
    print("The final acc is {!s}".format(true_acc))

    acc_list1, acc_list2 = selectsample(model, x_test, y_test, delta, iterate, fig_size)
    print("the select acc std is {!s}".format(np.mean(np.abs(acc_list1 - true_acc))))
    print("the random acc std is {!s}".format(np.mean(np.abs(acc_list2 - true_acc))))

    return np.array(acc_list1), np.array(acc_list2)

def get_cifar10_cn12(**kwargs):
    (X_train, Y_train), (X_test, Y_test) = cifar10.load_data()
    X_test = X_test.astype('float32')
    X_test = (X_test / 255.0) - (1.0 - 0.5)
    X_train = X_train.astype('float32')
    X_train = (X_train / 255.0) - (1.0 - 0.5)
    return X_test, Y_test, X_train, Y_train


def get_cifar10_vgg16(**kwargs):
    (X_train, Y_train), (X_test, Y_test) = cifar10.load_data()
    X_train = X_train.astype('float32')
    X_test = X_test.astype('float32')
    mean = np.mean(X_train, axis=(0, 1, 2, 3))
    std = np.std(X_train, axis=(0, 1, 2, 3))
    
    X_train = (X_train - mean) / (std + 1e-7)
    X_test = (X_test - mean) / (std + 1e-7)

    # Y_test = keras.utils.to_categorical(Y_test, 10)
    # Y_train = keras.utils.to_categorical(Y_train, 10)
    return X_test, Y_test, X_train, Y_train


def get_cifar100_vgg16(**kwargs):
    (X_train, Y_train), (X_test, Y_test) = cifar100.load_data()

    X_train = X_train.astype('float32')
    X_test = X_test.astype('float32')
    mean = np.mean(X_train, axis=(0, 1, 2, 3))
    std = np.std(X_train, axis=(0, 1, 2, 3))
    
    X_train = (X_train - mean) / (std + 1e-7)
    X_test = (X_test - mean) / (std + 1e-7)

    # Y_test = keras.utils.to_categorical(Y_test, 100)
    # Y_train = keras.utils.to_categorical(Y_train, 100)
    return X_test, Y_test, X_train, Y_train

def imagenet_preprocess(x):
    """
    Refer to keras.applications
    https://github.com/keras-team/keras/blob/df03bb5b1cc9fd297b0f19e08d916a4faedea267/keras/applications/imagenet_utils.py#L60
    x should be 'RGB' mode.
    """
    mean = [103.939, 116.779, 123.68]
    # change 'RGB' mode to 'BGR' mode.
    x = x[..., ::-1]
    x[..., 0] -= mean[0]
    x[..., 1] -= mean[1]
    x[..., 2] -= mean[2]
    return x

def get_imagenet(**kwargs):
    data_path = os.path.join(basedir, 'data', "imagenet.npz")
    data = np.load(data_path)
    X_test, Y_test = data['x_test'], data['y_test']
    exp_id = kwargs['exp_id']
    if exp_id == 'vgg19':
        X_test = vgg19.preprocess_input(X_test)
        # Y_test = keras.utils.to_categorical(Y_test, num_classes=1000)
    if exp_id == 'resnet50':
        X_test = resnet50.preprocess_input(X_test)
        # Y_test = keras.utils.to_categorical(Y_test, num_classes=1000)
    return X_test, Y_test

def  get_imagenet_5000(**kwargs):
    data_path = os.path.join(basedir, 'data', "imagenet_validation_5000.npz")
    data = np.load(data_path)
    X_test, Y_test = data['x_data'], data['y_data']
    # Y_test = keras.utils.to_categorical(Y_test, num_classes=1000)
    X_test = X_test.astype("float32")
    X_test = imagenet_preprocess(X_test)
    return X_test, Y_test

def get_combined_cifar10(**kwargs):
    # Load the CIFAR10 data.
    (x_train, y_train), (x_test, y_test) = cifar10.load_data()

    x_train = x_train.astype('float32')
    x_test = x_test.astype('float32')
    x_train /= 255
    x_test /= 255

    adv_data_path = os.path.join(basedir,'cifar10_vgg16_jsma_full.npz')
    adv_data = np.load(adv_data_path)

    x_adv = adv_data["inputs"]
    y_adv = adv_data["labels"]
    y_adv = y_adv.squeeze()

    x_test = x_test[:5000]
    y_test = y_test[:5000]
    y_test = y_test.reshape(5000)
    
    x_test = np.concatenate((x_test,x_adv[:5000]),axis=0)
    y_test = np.concatenate((y_test,y_adv[:5000]),axis=0)

    return x_test, y_test

def get_combined_cifar100(**kwargs):
    (x_train, y_train), (x_test, y_test) = cifar100.load_data()

    x_train = x_train.astype('float32')
    x_test = x_test.astype('float32')
    x_train /= 255
    x_test /= 255

    means = [0.5070746, 0.48654896, 0.44091788]

    x_train = subtract_mean(x_train, means)
    x_test = subtract_mean(x_test, means)

    adv_data_path = os.path.join(basedir,'cifar100_resnet32_fgsm_full.npz')
    adv_data = np.load(adv_data_path)

    x_adv = adv_data["inputs"]
    y_adv = adv_data["labels"]
    y_adv = y_adv.squeeze()

    x_test = x_test[:5000]
    y_test = y_test[:5000]
    y_test = y_test.reshape(5000)

    x_test = np.concatenate((x_test,x_adv[:5000]),axis=0)
    y_test = np.concatenate((y_test,y_adv[:5000]),axis=0)

    return x_test, y_test

def get_data(exp_id):
    exp_model_dict = {'combined_cifar10': get_combined_cifar10,
                      'combined_cifar100': get_combined_cifar100,
                      'cn12': get_cifar10_cn12,
                      'cifar10_vgg16': get_cifar10_vgg16,
                      'cifar100_vgg16': get_cifar100_vgg16}
    return exp_model_dict[exp_id](exp_id=exp_id)


def get_model(exp_id):
    basedir = os.path.abspath(os.path.dirname(__file__))

    exp_model_dict = {'combined_cifar10': 'model/cifar10_vgg16.h5',
                      'combined_cifar100': 'model/cifar100_resnet32.h5',
                      'cn12': 'model/cifar10_cn12.h5',
                      'cifar10_vgg16': 'model/cifar10vgg.h5',
                      'cifar100_vgg16': 'model/cifar100vgg.h5'}

    if exp_id == 'vgg19' or exp_id == 'vgg19_5000':
        my_model = vgg19.VGG19(weights='imagenet')
        adam = keras.optimizers.Adagrad(lr=0.01, epsilon=None, decay=0.0)
        my_model.compile(loss='categorical_crossentropy', optimizer=adam, metrics=['accuracy'])
    elif exp_id == 'resnet50' or exp_id == 'resnet50_5000':
        my_model = resnet50.ResNet50(weights='imagenet')
        adam = keras.optimizers.Adagrad(lr=0.01, epsilon=None, decay=0.0)
        my_model.compile(loss='categorical_crossentropy', optimizer=adam, metrics=['accuracy'])
    elif exp_id in exp_model_dict.keys():
        my_model = keras.models.load_model(os.path.join(basedir, exp_model_dict[exp_id]))
    else:
        raise Exception("no such dataset found: {}".format(exp_id))

    return my_model


def get_acc(exp_id):
    acc_dict = {'combined_cifar10': 0.437,
                'combined_cifar10': 0.3504,
                'cn12': 0.8064,
                'cifar10_vgg16': 0.9375,
                'cifar100_vgg16': 0.6944}
    return acc_dict[exp_id]

def get_fig_size(exp_id):
    if exp_id in ['vgg19', 'vgg19_5000', 'resnet50', 'resnet50_5000']:
        return 224
    else:
        return 32


if __name__ == '__main__':

    basedir = os.path.abspath(os.path.dirname(__file__))

    """Parser of command args"""
    parse = argparse.ArgumentParser()
    parse.add_argument("--exp_id", type=str, help="exp_identifiers")
    # parse.add_argument("--pre_sample_num", type=int, help='Pre_sampling number')
    parse.add_argument("--random_seed", type=int, help='Random seed')

    console_flags, unparsed = parse.parse_known_args(sys.argv[1:])

    exp_id = console_flags.exp_id
    random_seed = console_flags.random_seed

    acc = get_acc(exp_id)

    model = get_model(exp_id=exp_id)
    x_test, y_test, x_train, y_train = get_data(exp_id=exp_id)
    # x_test, y_test = get_data(exp_id=exp_id)
    print(model.summary())

    start = datetime.datetime.now()

    fig_size = get_fig_size(exp_id=exp_id)

    test = x_test
    output = model.predict(test)
    max_output = np.max(output, axis=1)
    index = np.argsort(max_output)
    # devide into 3 strata
    index1 = index[0:int(x_test.shape[0] * 0.1)]
    index2 = index[int(x_test.shape[0] * 0.1):int(x_test.shape[0] * 0.2)]
    index3 = index[int(x_test.shape[0] * 0.2):]

    label = y_test[index1]
    pred = np.argmax(output[index1], axis=1)
    s1 = np.std(pred == label)
    print("strata1_Sh", s1)
    acc_1 = np.sum(pred == label) / len(pred)
    print("strata1_acc", acc_1)

    label = y_test[index2]
    pred = np.argmax(output[index2], axis=1)
    s2 = np.std(pred == label)
    print("strata2_Sh", s2)
    acc_2 = np.sum(pred == label) / len(pred)
    print("strata2_acc", acc_2)

    label = y_test[index3]
    pred = np.argmax(output[index3], axis=1)
    s3 = np.std(pred == label)
    print("strata3_Sh", s3)
    acc_3 = np.sum(pred == label) / len(pred)
    print("strata3_acc", acc_3)

    label = y_test[index1]
    pred = np.argmax(output[index1], axis=1)
    s1 = np.std(pred == label)
    print("strata1_Sh", s1)
    acc_1 = np.sum(pred == label) / len(pred)
    print("strata1_acc", acc_1)

    total_adaptive_sampling = 30
    np.random.seed(random_seed)

    arr = np.random.permutation(index1)
    random_index = arr[0:10]
    random_index = random_index.astype('int')
    index1_adaptive = random_index
    print(index1_adaptive)
    label = y_test[index1_adaptive]
    print(output[index1_adaptive])
    pred = np.argmax(output[index1_adaptive], axis=1)
    print(label)
    print(pred)
    s1 = np.std(pred == label)
    print("strata1_Sh", s1)
    acc_1 = np.sum(pred == label) / len(pred)
    print("strata1_acc", acc_1)

    arr = np.random.permutation(index2)
    random_index = arr[0:10]
    random_index = random_index.astype('int')
    index2_adaptive = random_index
    print(index2_adaptive)
    label = y_test[index2_adaptive]
    pred = np.argmax(output[index2_adaptive], axis=1)
    s2 = np.std(pred == label)
    print("strata2_Sh", s2)
    acc_2 = np.sum(pred == label) / len(pred)
    print("strata2_acc", acc_2)

    arr = np.random.permutation(index3)
    random_index = arr[0:10]
    random_index = random_index.astype('int')
    index3_adaptive = random_index
    print(index3_adaptive)
    label = y_test[index3_adaptive]
    pred = np.argmax(output[index3_adaptive], axis=1)
    s3 = np.std(pred == label)
    print("strata3_Sh", s3)
    acc_3 = np.sum(pred == label) / len(pred)
    print("strata3_acc", acc_3)

    total_ws = len(index1) * s1 + len(index2) * s2 + len(index3) * s3
    w = [len(index1) * s1 / total_ws, len(index2) * s2 / total_ws, len(index3) * s3 / total_ws]
    print(w)
    print(sum(w))

    acc_select = []
    acc_random = []

    for k in range(50):
        print("the {} exp".format(k))
        acc1, acc2 = experiments(delta=10, iterate=17, fig_size=fig_size)

        acc1 = np.square(np.array(acc1) - acc)
        acc2 = np.square(np.array(acc2) - acc)

        acc_select.append(acc1)
        acc_random.append(acc2)

        elapsed = (datetime.datetime.now() - start)
        print("Time used: ", elapsed)

    acc_select = np.array(acc_select)
    acc_random = np.array(acc_random)

    mse_select = np.sqrt(acc_select.mean(axis=0))
    mse_random = np.sqrt(acc_random.mean(axis=0))

    # save the result
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    row_num = 7
    sheet.cell(row=row_num, column=1).value = "SSOA-M-P"
    for i in range(len(mse_select)):
        sheet.cell(row=row_num, column=i + 2).value = mse_select[i]
    row_num += 1
    for i in range(len(mse_random)):
        sheet.cell(row=row_num, column=i + 2).value = mse_random[i]

    workbook.save(os.path.join(basedir, "results", "{}-ssoamp-randomseed{}.xlsx".format(exp_id, random_seed)))

    elapsed = (datetime.datetime.now() - start)
    print("Time used: ", elapsed)