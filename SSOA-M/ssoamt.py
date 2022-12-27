# -*- coding: UTF-8 -*-
import sys
import datetime
import os
import numpy as np
import tensorflow.keras.backend as K
from collections import defaultdict
import tensorflow.keras as keras
from sklearn.metrics import *
from tensorflow.keras.datasets import cifar10, cifar100
from tensorflow.keras.models import load_model
import argparse
from tensorflow.keras.applications import vgg19, resnet50
import openpyxl

def get_score(x_test, y_test, model):
    score = model.evaluate(x_test, y_test, verbose=0)
    print('Test loss:', score[0])
    print('Test accuracy:', score[1])
    return score

def selectsample(model, x_test, y_test, delta, iterate):
    test = x_test
    output = model.predict(test)
    max_output = np.max(output, axis=1)
    index = np.argsort(max_output)
    # devide into 3 strata
    index1 = index[0:int(x_test.shape[0] * 0.1)]
    index2 = index[int(x_test.shape[0] * 0.1):int(x_test.shape[0] * 0.2)]
    index3 = index[int(x_test.shape[0] * 0.2):]

    train = x_train
    output_train = model.predict(train)
    max_output_train = np.max(output_train, axis=1)
    index1_train = []
    index2_train = []
    index3_train = []
    for i in range(len(max_output_train)):
        if max_output_train[i] >= max_output[index1[0]] and max_output_train[i] <= max_output[index1[-1]]:
            index1_train.append(i)
        elif max_output_train[i] >= max_output[index2[0]] and max_output_train[i] <= max_output[index2[-1]]:
            index2_train.append(i)
        elif max_output_train[i] >= max_output[index3[0]] and max_output_train[i] <= max_output[index3[-1]]:
            index3_train.append(i)

    print(len(index1_train))
    print(len(index2_train))
    print(len(index3_train))

    label = y_train[index1_train]
    pred = np.argmax(output_train[index1_train], axis=1)
    s1 = np.std(pred == label)
    print("strata1_Sh", s1)
    acc_1 = np.sum(pred == label) / len(pred)
    print("strata1_acc", acc_1)

    label = y_train[index2_train]
    pred = np.argmax(output_train[index2_train], axis=1)
    s2 = np.std(pred == label)
    print("strata2_Sh", s2)
    acc_2 = np.sum(pred == label) / len(pred)
    print("strata2_acc", acc_2)

    label = y_train[index3_train]
    pred = np.argmax(output_train[index3_train], axis=1)
    s3 = np.std(pred == label)
    print("strata3_Sh", s3)
    acc_3 = np.sum(pred == label) / len(pred)
    print("strata3_acc", acc_3)

    total_ws = len(index1_train) * s1 + len(index2_train) * s2 + len(index3_train) * s3
    w = [len(index1_train) * s1 / total_ws, len(index2_train) * s2 / total_ws, len(index3_train) * s3 / total_ws]
    print(w)
    print(sum(w))

    acc_list1 = []
    acc_list2 = []

    select_num = 50
    for i in range(iterate):
        if w[0] == 0:
            acc_1 = 1
        else:
            arr = np.random.permutation(index1.shape[0])
            temp_index = arr[0:int(select_num * w[0])]
            statra_index1 = index1[temp_index]
            statra_index1 = statra_index1.astype('int')
            label = y_test[statra_index1]
            orig_sample = x_test[statra_index1]
            orig_sample = orig_sample.reshape(-1, 32, 32, 3)
            pred = np.argmax(model.predict(orig_sample), axis=1)
            acc_1 = np.sum(pred == label) / orig_sample.shape[0]

        if w[1] == 0:
            acc_2 = 1
        else:
            arr = np.random.permutation(index2.shape[0])
            temp_index = arr[0:int(select_num * w[1])]
            statra_index2 = index2[temp_index]
            statra_index2 = statra_index2.astype('int')
            label = y_test[statra_index2]
            orig_sample = x_test[statra_index2]
            orig_sample = orig_sample.reshape(-1, 32, 32, 3)
            pred = np.argmax(model.predict(orig_sample), axis=1)
            acc_2 = np.sum(pred == label) / orig_sample.shape[0]

        if w[2] == 0:
            acc_3 = 1
        else:
            arr = np.random.permutation(index3.shape[0])
            temp_index = arr[0:int(select_num * w[2])]
            statra_index3 = index3[temp_index]
            statra_index3 = statra_index3.astype('int')
            label = y_test[statra_index3]
            orig_sample = x_test[statra_index3]
            orig_sample = orig_sample.reshape(-1, 32, 32, 3)
            pred = np.argmax(model.predict(orig_sample), axis=1)
            acc_3 = np.sum(pred == label) / orig_sample.shape[0]

        arr = np.random.permutation(x_test.shape[0])
        random_index = arr[0:select_num]
        random_index = random_index.astype('int')

        acc1 = 0.1 * acc_1 + 0.1 * acc_2 + 0.8 * acc_3
        acc_list1.append(acc1)

        label = y_test[random_index]
        orig_sample = x_test[random_index]
        orig_sample = orig_sample.reshape(-1, 32, 32, 3)
        pred = np.argmax(model.predict(orig_sample), axis=1)
        acc2 = np.sum(pred == label) / orig_sample.shape[0]
        acc_list2.append(acc2)

        print("numuber of samples is {!s}, select acc is {!s}, random acc is {!s}".format(
            orig_sample.shape[0], acc1, acc2))

        select_num = select_num + delta

    return acc_list1, acc_list2


def experiments(delta, iterate):
    pred = np.argmax(model.predict(x_test), axis=1)
    true_acc = np.sum(pred == y_test) / x_test.shape[0]
    print("The final acc is {!s}".format(true_acc))

    acc_list1, acc_list2 = selectsample(
        model, x_test, y_test, delta, iterate)
    print("the select acc std is {!s}".format(
        np.mean(np.abs(acc_list1 - true_acc))))
    print("the random acc std is {!s}".format(
        np.mean(np.abs(acc_list2 - true_acc))))

    return np.array(acc_list1), np.array(acc_list2)

def get_cifar10_cn12(**kwargs):
    (X_train, Y_train), (X_test, Y_test) = cifar10.load_data()
    X_test = X_test.astype('float32')
    X_test = (X_test / 255.0) - (1.0 - 0.5)
    # Y_test = keras.utils.to_categorical(Y_test, 10)

    X_train = X_train.astype('float32')
    X_train = (X_train / 255.0) - (1.0 - 0.5)
    # Y_train = keras.utils.to_categorical(Y_train, 10)
    return X_test, Y_test, X_train, Y_train


def get_cifar10_vgg16(**kwargs):
    (X_train, Y_train), (X_test, Y_test) = cifar10.load_data()

    X_train = X_train.astype('float32')
    X_test = X_test.astype('float32')
    mean = np.mean(X_train, axis=(0, 1, 2, 3))
    std = np.std(X_train, axis=(0, 1, 2, 3))
    
    X_train = (X_train - mean) / (std + 1e-7)
    X_test = (X_test - mean) / (std + 1e-7)

    # Y_test = keras.utils.to_categorical(Y_test, 10)
    # Y_train = keras.utils.to_categorical(Y_train, 10)
    return X_test, Y_test, X_train, Y_train


def get_cifar100_vgg16(**kwargs):
    (X_train, Y_train), (X_test, Y_test) = cifar100.load_data()

    X_train = X_train.astype('float32')
    X_test = X_test.astype('float32')
    mean = np.mean(X_train, axis=(0, 1, 2, 3))
    std = np.std(X_train, axis=(0, 1, 2, 3))
    
    X_train = (X_train - mean) / (std + 1e-7)
    X_test = (X_test - mean) / (std + 1e-7)

    # Y_test = keras.utils.to_categorical(Y_test, 100)
    # Y_train = keras.utils.to_categorical(Y_train, 100)
    return X_test, Y_test, X_train, Y_train

def get_imagenet(**kwargs):
    data_path = os.path.join(basedir, 'data', "imagenet.npz")
    data = np.load(data_path)
    X_test, Y_test = data['x_test'], data['y_test']
    # X_train, Y_train = data['x_train'], data['y_train']
    exp_id = kwargs['exp_id']
    if exp_id == 'vgg19':
        X_test = vgg19.preprocess_input(X_test)
        Y_test = keras.utils.to_categorical(Y_test, num_classes=1000)
    if exp_id == 'resnet50':
        X_test = resnet50.preprocess_input(X_test)
        Y_test = keras.utils.to_categorical(Y_test, num_classes=1000)
    return X_test, Y_test

def get_data(exp_id):
    exp_model_dict = {'cn12': get_cifar10_cn12,
                      'cifar10_vgg16': get_cifar10_vgg16,
                      'cifar100_vgg16': get_cifar100_vgg16}
    return exp_model_dict[exp_id](exp_id=exp_id)


def get_model(exp_id):
    basedir = os.path.abspath(os.path.dirname(__file__))

    exp_model_dict = {'cn12': 'model/cifar10_cn12.h5',
                      'cifar10_vgg16': 'model/cifar10vgg.h5',
                      'cifar100_vgg16': 'model/cifar100vgg.h5'}

    if exp_id == 'vgg19':
        my_model = vgg19.VGG19(weights='imagenet')
        adam = keras.optimizers.Adagrad(lr=0.01, epsilon=None, decay=0.0)
        my_model.compile(loss='categorical_crossentropy', optimizer=adam, metrics=['accuracy'])
    elif exp_id == 'resnet50':
        my_model = resnet50.ResNet50(weights='imagenet')
        adam = keras.optimizers.Adagrad(lr=0.01, epsilon=None, decay=0.0)
        my_model.compile(loss='categorical_crossentropy', optimizer=adam, metrics=['accuracy'])
    elif exp_id in exp_model_dict.keys():
        my_model = keras.models.load_model(os.path.join(basedir, exp_model_dict[exp_id]))
    else:
        raise Exception("no such dataset found: {}".format(exp_id))

    return my_model


def get_acc(exp_id):
    acc_dict = {'cn12': 0.8064,
                'cifar10_vgg16': 0.9375,
                'cifar100_vgg16': 0.6944}
    return acc_dict[exp_id]


if __name__ == '__main__':

    basedir = os.path.abspath(os.path.dirname(__file__))

    """Parser of command args"""
    parse = argparse.ArgumentParser()
    parse.add_argument("--exp_id", type=str, help="exp_identifiers")

    console_flags, unparsed = parse.parse_known_args(sys.argv[1:])

    exp_id = console_flags.exp_id

    acc = get_acc(exp_id)

    model = get_model(exp_id=exp_id)
    x_test, y_test, x_train, y_train = get_data(exp_id=exp_id)
    print(model.summary())
    print(x_test.shape)
    print(x_train.shape)

    y_test = y_test.squeeze()
    y_train = y_train.squeeze()

    # get_score(x_test,y_test,my_model)
    # get_score(x_train,y_train,my_model)

    start = datetime.datetime.now()

    acc_select = []
    acc_random = []

    for k in range(50):
        print("the {} exp".format(k))
        acc1, acc2 = experiments(delta=10, iterate=97)

        acc1 = np.square(np.array(acc1) - acc)
        acc2 = np.square(np.array(acc2) - acc)

        acc_select.append(acc1)
        acc_random.append(acc2)

        elapsed = (datetime.datetime.now() - start)
        print("Time used: ", elapsed)

    acc_select = np.array(acc_select)
    acc_random = np.array(acc_random)

    mse_select = np.sqrt(acc_select.mean(axis=0))
    mse_random = np.sqrt(acc_random.mean(axis=0))

    # save the result
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    row_num = 6
    sheet.cell(row=row_num, column=1).value = "SSOA-M-T"
    for i in range(len(mse_select)):
        sheet.cell(row=row_num, column=i + 2).value = mse_select[i]
    row_num += 1
    for i in range(len(mse_random)):
        sheet.cell(row=row_num, column=i + 2).value = mse_random[i]

    workbook.save(os.path.join(basedir, "results", "{}-ssoamt.xlsx".format(exp_id)))

    elapsed = (datetime.datetime.now() - start)
    print("Time used: ", elapsed)