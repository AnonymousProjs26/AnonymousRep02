# -*- coding: UTF-8 -*-
import sys
import datetime
import os
import numpy as np
import tensorflow.keras.backend as K
from collections import defaultdict
import matplotlib.pyplot as plt
import tensorflow.keras as keras
from sklearn.metrics import *
import math
from tensorflow.keras.datasets import cifar10, cifar100
from tensorflow.keras.models import load_model
import argparse
from tensorflow.keras.applications import vgg19, resnet50
import openpyxl

def selectsample(model, X_test_preds, Y_test, delta, iterate, fig_size):
    global index3, index1, index2, index1_adaptive, index2_adaptive, index3_adaptive, total_adaptive_sampling, w
    acc_list1 = []
    acc_list2 = []

    select_num = 50 - total_adaptive_sampling
    for i in range(iterate):
        if w[0] == 0 or math.ceil(select_num * w[0]) < 1:
            acc_1 = 1
        else:
            index1 = np.setdiff1d(index1, index1_adaptive, True)
            arr = np.random.permutation(index1.shape[0])
            temp_index = arr[0:math.ceil(select_num * w[0])]
            statra_index1 = index1[temp_index]
            statra_index1 = statra_index1.astype('int')

            statra_index1 = np.union1d(statra_index1, index1_adaptive)
            print("0: ", len(statra_index1))
            label = Y_test[statra_index1]
            orig_sample = X_test_preds[statra_index1]
            # orig_sample = orig_sample.reshape(-1, fig_size, fig_size, 3)
            pred = np.argmax(orig_sample, axis=1)
            acc_1 = np.sum(pred == label) / orig_sample.shape[0]


        if w[1] == 0 or math.ceil(select_num * w[1]) < 1:
            acc_2 = 1
        else:
            index2 = np.setdiff1d(index2, index2_adaptive, True)
            arr = np.random.permutation(index2.shape[0])
            temp_index = arr[0:math.ceil(select_num * w[1])]
            statra_index2 = index2[temp_index]
            statra_index2 = statra_index2.astype('int')

            statra_index2 = np.union1d(statra_index2, index2_adaptive)
            print("1: ", len(statra_index2))
            label = Y_test[statra_index2]
            orig_sample = X_test_preds[statra_index2]
            # orig_sample = orig_sample.reshape(-1, fig_size, fig_size, 3)
            pred = np.argmax(orig_sample, axis=1)
            acc_2 = np.sum(pred == label) / orig_sample.shape[0]

        if w[2] == 0 or math.ceil(select_num * w[2]) < 1:
            acc_3 = 1
        else:
            index3 = np.setdiff1d(index3, index3_adaptive, True)
            arr = np.random.permutation(index3.shape[0])
            temp_index = arr[0:math.ceil(select_num * w[2])]
            statra_index3 = index3[temp_index]
            statra_index3 = statra_index3.astype('int')

            statra_index3 = np.union1d(statra_index3, index3_adaptive)
            print("2: ", len(statra_index3))
            label = Y_test[statra_index3]
            orig_sample = X_test_preds[statra_index3]
            # orig_sample = orig_sample.reshape(-1, fig_size, fig_size, 3)
            pred = np.argmax(orig_sample, axis=1)
            acc_3 = np.sum(pred == label) / orig_sample.shape[0]

        arr = np.random.permutation(X_test_preds.shape[0])
        random_index = arr[0:select_num+total_adaptive_sampling]
        random_index = random_index.astype('int')

        acc1 = 0.1 * acc_1 + 0.1 * acc_2 + 0.8 * acc_3
        acc_list1.append(acc1)

        label = Y_test[random_index]
        orig_sample = X_test_preds[random_index]
        # orig_sample = orig_sample.reshape(-1, fig_size, fig_size, 3)
        pred = np.argmax(orig_sample, axis=1)
        acc2 = np.sum(pred == label) / orig_sample.shape[0]
        acc_list2.append(acc2)

        print("numuber of samples is {!s}, select acc is {!s}, random acc is {!s}".format(
            orig_sample.shape[0], acc1, acc2))

        select_num = select_num + delta

    return acc_list1, acc_list2


def experiments(delta, iterate, fig_size):
    pred = np.argmax(X_test_preds, axis=1)
    true_acc = np.sum(pred == Y_test) / Y_test.shape[0]
    print("The final acc is {!s}".format(true_acc))

    acc_list1, acc_list2 = selectsample(model, X_test_preds, Y_test, delta, iterate, fig_size)
    print("the select acc std is {!s}".format(np.mean(np.abs(acc_list1 - true_acc))))
    print("the random acc std is {!s}".format(np.mean(np.abs(acc_list2 - true_acc))))

    return np.array(acc_list1), np.array(acc_list2)

def load_imagenet_result(model_name):
    filename1 = f"./imagenet_tools/dataset/{model_name}_train_50000.npz"
    train_result = np.load(filename1, allow_pickle=True)
    train_preds, train_labels = train_result["preds"], train_result["labels"]

    filename2 = f"./imagenet_tools/dataset/{model_name}_val_5000.npz"
    test_result = np.load(filename2, allow_pickle=True)
    test_preds, test_labels = test_result["preds"], test_result["labels"]
    return train_preds, train_labels, test_preds, test_labels


def get_model(exp_id):
    if exp_id == 'vgg19' or exp_id == 'vgg19_5000':
        my_model = vgg19.VGG19(weights='imagenet')
        adam = keras.optimizers.Adagrad(lr=0.01, epsilon=None, decay=0.0)
        my_model.compile(loss='categorical_crossentropy', optimizer=adam, metrics=['accuracy'])
    elif exp_id == 'resnet50' or exp_id == 'resnet50_5000':
        my_model = resnet50.ResNet50(weights='imagenet')
        adam = keras.optimizers.Adagrad(lr=0.01, epsilon=None, decay=0.0)
        my_model.compile(loss='categorical_crossentropy', optimizer=adam, metrics=['accuracy'])
    else:
        raise Exception("no such dataset found: {}".format(exp_id))

    return my_model


def get_acc(exp_id):
    acc_dict = {'vgg19': 0.7092,
                'resnet50': 0.7496,
                'vgg19_5000': 0.7146000266075134,
                'resnet50_5000': 0.7513999938964844}
    return acc_dict[exp_id]

def get_fig_size(exp_id):
    if exp_id in ['vgg19', 'vgg19_5000', 'resnet50', 'resnet50_5000']:
        return 224
    else:
        return 32


if __name__ == '__main__':

    basedir = os.path.abspath(os.path.dirname(__file__))

    """Parser of command args"""
    parse = argparse.ArgumentParser()
    parse.add_argument("--exp_id", type=str, help="exp_identifiers")
    # parse.add_argument("--pre_sample_num", type=int, help='Pre_sampling number')
    parse.add_argument("--random_seed", type=int, help='Random seed')

    console_flags, unparsed = parse.parse_known_args(sys.argv[1:])

    exp_id = console_flags.exp_id
    random_seed = console_flags.random_seed

    acc = get_acc(exp_id)

    model = get_model(exp_id=exp_id)
    X_train_preds, Y_train, X_test_preds, Y_test = load_imagenet_result(model_name=exp_id)
    train_pred_labels = np.argmax(X_train_preds, axis=1)
    test_pred_labels = np.argmax(X_test_preds, axis=1)
    print(model.summary())

    start = datetime.datetime.now()

    fig_size = get_fig_size(exp_id=exp_id)

    output = X_test_preds
    max_output = np.max(output, axis=1)
    index = np.argsort(max_output)
    # devide into 3 strata
    index1 = index[0:int(X_test_preds.shape[0] * 0.1)]
    index2 = index[int(X_test_preds.shape[0] * 0.1):int(X_test_preds.shape[0] * 0.2)]
    index3 = index[int(X_test_preds.shape[0] * 0.2):]

    total_adaptive_sampling = 30
    np.random.seed(random_seed)

    arr = np.random.permutation(index1)
    random_index = arr[0:10]
    random_index = random_index.astype('int')
    index1_adaptive = random_index
    print(index1_adaptive)
    label = Y_test[index1_adaptive]
    print(output[index1_adaptive])
    pred = np.argmax(output[index1_adaptive], axis=1)
    print(label)
    print(pred)
    s1 = np.std(pred == label)
    print("strata1_Sh", s1)
    acc_1 = np.sum(pred == label) / len(pred)
    print("strata1_acc", acc_1)

    arr = np.random.permutation(index2)
    random_index = arr[0:10]
    random_index = random_index.astype('int')
    index2_adaptive = random_index
    print(index2_adaptive)
    label = Y_test[index2_adaptive]
    pred = np.argmax(output[index2_adaptive], axis=1)
    s2 = np.std(pred == label)
    print("strata2_Sh", s2)
    acc_2 = np.sum(pred == label) / len(pred)
    print("strata2_acc", acc_2)

    arr = np.random.permutation(index3)
    random_index = arr[0:10]
    random_index = random_index.astype('int')
    index3_adaptive = random_index
    print(index3_adaptive)
    label = Y_test[index3_adaptive]
    pred = np.argmax(output[index3_adaptive], axis=1)
    s3 = np.std(pred == label)
    print("strata3_Sh", s3)
    acc_3 = np.sum(pred == label) / len(pred)
    print("strata3_acc", acc_3)

    total_ws = len(index1) * s1 + len(index2) * s2 + len(index3) * s3
    w = [len(index1) * s1 / total_ws, len(index2) * s2 / total_ws, len(index3) * s3 / total_ws]
    print(w)
    print(sum(w))

    acc_select = []
    acc_random = []

    for k in range(50):
        print("the {} exp".format(k))
        acc1, acc2 = experiments(delta=10, iterate=98, fig_size=fig_size)

        acc1 = np.square(np.array(acc1) - acc)
        acc2 = np.square(np.array(acc2) - acc)

        acc_select.append(acc1)
        acc_random.append(acc2)

        elapsed = (datetime.datetime.now() - start)
        print("Time used: ", elapsed)

    acc_select = np.array(acc_select)
    acc_random = np.array(acc_random)

    mse_select = np.sqrt(acc_select.mean(axis=0))
    mse_random = np.sqrt(acc_random.mean(axis=0))

    # save the result
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    row_num = 7
    sheet.cell(row=row_num, column=1).value = "SSOA-M-P"
    for i in range(len(mse_select)):
        sheet.cell(row=row_num, column=i + 2).value = mse_select[i]
    row_num += 1
    for i in range(len(mse_random)):
        sheet.cell(row=row_num, column=i + 2).value = mse_random[i]

    workbook.save(os.path.join(basedir, "results", "{}-ssoamp-randomseed{}.xlsx".format(exp_id, random_seed)))

    elapsed = (datetime.datetime.now() - start)
    print("Time used: ", elapsed)